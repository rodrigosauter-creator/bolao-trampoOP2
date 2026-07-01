import openpyxl
import json

ARQUIVO_EXCEL = "Copa do Mundo 2026.xlsx"
ARQUIVO_JSON = "dados.json"

# =========================
# Abrir planilha
# =========================

wb = openpyxl.load_workbook(
    ARQUIVO_EXCEL,
    data_only=True
)

# =========================
# CLASSIFICAÇÃO
# =========================

ws = wb["Classificação"]

classificacao = []

for row in ws.iter_rows(
    min_row=3,
    max_col=8,
    values_only=True
):

    posicao = row[4]
    nome = row[5]
    pontos = row[6]
    acertos = row[7]

    if nome is None:
        continue

    classificacao.append({
        "posicao": int(posicao) if posicao is not None else 0,
        "nome": str(nome),
        "pontos": int(pontos) if pontos is not None else 0,
        "acertos": int(acertos) if acertos is not None else 0
    })

def tratar_penaltis(valor, jogo):

    if int(jogo) < 73:
        return None

    if valor is None or str(valor).strip() == "":
        return "x"

    return str(valor).strip()


def tratar_palpite_certo(valor):

    if valor is None:
        return 0

    valor_texto = str(valor).strip().lower()

    if valor_texto in ["1", "true", "verdadeiro", "sim"]:
        return 1

    return 0
    
# =========================
# EVOLUÇÃO
# =========================

ws = wb["Pontuação por rodada"]

cabecalho = list(
    next(
        ws.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True
        )
    )
)

participantes = []

for nome in cabecalho[1:8]:

    if nome is not None:
        participantes.append(
            str(nome).strip()
        )

evolucao = {}

for participante in participantes:
    evolucao[participante] = []

for row in ws.iter_rows(
    min_row=2,
    values_only=True
):

    jogo = row[0]

    if jogo is None:
        continue

    for i, participante in enumerate(
        participantes,
        start=1
    ):

        valor = row[i]

        evolucao[participante].append({
            "jogo": int(jogo),
            "pontos": valor if valor is not None else 0
        })
        
# =========================
# JOGOS
# =========================

ws = wb["Tabela de Jogos"]

jogos = []
jogos_realizados = 0

for row in ws.iter_rows(
    min_row=3,
    values_only=True
):

    jogo = row[0]

    if jogo is None:
        continue

    selecao_a = row[1]
    gols_a = row[2]
    gols_b = row[4]
    selecao_b = row[5]
    vencedor = row[6]
    penaltis = tratar_penaltis(row[7], jogo)
    
    realizado = (
        gols_a is not None
        and gols_b is not None
    )

    

    if realizado:
        jogos_realizados += 1

    jogos.append({
        "jogo": int(jogo),
        "selecao_a": selecao_a,
        "gols_a": gols_a,
        "gols_b": gols_b,
        "selecao_b": selecao_b,
        "vencedor": vencedor,
        "penaltis": penaltis,
        "realizado": realizado
    })

# =========================
# PALPITES DOS APOSTADORES
# =========================

abas_ignoradas = [
    "Tabela de Jogos",
    "Classificação",
    "Pontuação por rodada"
]

apostadores = {}

for nome_aba in wb.sheetnames:

    if nome_aba in abas_ignoradas:
        continue

    ws = wb[nome_aba]

    palpites = []

    total = 0
    acertos = 0

    for row in ws.iter_rows(
        min_row=3,
        values_only=True
    ):

        jogo = row[8]

        if jogo is None:
            continue

        pontos = row[22]

        if row[23] is not None:
            total = row[23]

        if row[24] is not None:
            acertos = row[24]


    print(
    nome_aba,
    jogo,
    row[21]
)
    
    palpites.append({
    "jogo": int(jogo),
    "selecao_a": row[9],
    "gols_a": row[10],
    "gols_b": row[12],
    "selecao_b": row[13],
    "vencedor": row[14],
    "penaltis": tratar_penaltis(row[15], jogo),
    "palpite_certo": tratar_palpite_certo(row[21]),
    "pontos": pontos if pontos is not None else 0
})

    apostadores[nome_aba] = {
        "nome": nome_aba,
        "total": total,
        "acertos": acertos,
        "palpites": palpites
    }

# =========================
# RANKING POR JOGO
# =========================

ws = wb["Pontuação por rodada"]

ranking_por_rodada = []

linha = 2

while True:

    jogo = ws.cell(
        row=linha,
        column=10
    ).value

    if jogo is None:
        break

    ranking = []

    for col in range(11, 18):

        nome = ws.cell(
            row=1,
            column=col
        ).value

        pontos = ws.cell(
            row=linha,
            column=col
        ).value

        ranking.append({
            "nome": nome,
            "pontos": pontos if pontos is not None else 0
        })

    ranking.sort(
        key=lambda x: x["pontos"],
        reverse=True
    )

    ranking_por_rodada.append({
        "jogo": int(jogo),
        "vencedor_rodada": ranking[0]["nome"],
        "maior_pontuacao": ranking[0]["pontos"],
        "ranking": ranking
    })

    linha += 1
    
# =========================
# HISTÓRICO DE LIDERANÇA
# =========================

lideres = []

for indice_jogo in range(len(jogos)):

    ranking = []

    for participante in participantes:

        if indice_jogo >= len(
            evolucao[participante]
        ):
            continue

        pontos = (
            evolucao[participante]
            [indice_jogo]["pontos"]
        )

        ranking.append({
            "nome": participante,
            "pontos": pontos
        })

    ranking.sort(
        key=lambda x: x["pontos"],
        reverse=True
    )

    if len(ranking) > 0:

        if indice_jogo == 0:

            lider = "-"
            pontos = 0

        else:

            lider = ranking[0]["nome"]
            pontos = ranking[0]["pontos"]

        lideres.append({
            "jogo": indice_jogo,
            "lider": lider,
            "pontos": pontos
        })
        
# =========================
# EXPORTAÇÃO
# =========================

dados = {
    "classificacao": classificacao,
    "evolucao": evolucao,
    "jogos_realizados": jogos_realizados,
    "jogos": jogos,
    "apostadores": apostadores,
    "ranking_por_rodada": ranking_por_rodada,
    "lideres": lideres
}

with open(
    ARQUIVO_JSON,
    "w",
    encoding="utf-8"
) as f:

    json.dump(
        dados,
        f,
        ensure_ascii=False,
        indent=2
    )

wb.close()

print(
    f"dados.json gerado com sucesso. "
    f"Jogos realizados: {jogos_realizados}"
)
